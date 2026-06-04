"""
PitchGuard — Generate Player Stats Links Excel
===============================================
Creates an Excel with TWO sheets:
  Sheet 1 "TM Links"   — serial no, player name, club, TM stats URL (clickable)
  Sheet 2 "Paste Data" — empty table where you paste copied TM data

Usage:
    python generate_stats_links.py \
        --input  data/raw/player_bios.csv \
        --output data/raw/player_stats_links.xlsx
"""

import argparse
import re
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path


def extract_slug(profile_url: str) -> str:
    m = re.search(r"transfermarkt\.[a-z.]+/([^/]+)/profil", profile_url or "")
    return m.group(1) if m else ""


def build_stats_url(slug: str, tm_id: str) -> str:
    if not slug or not tm_id:
        return ""
    return (
        f"https://www.transfermarkt.com/{slug}"
        f"/leistungsdatendetails/spieler/{tm_id}/saison/ges/plus/0"
    )


# ── Style helpers ─────────────────────────────────────────────────────────────
def header_style(cell, bg: str = "1F4E79") -> None:
    cell.font = Font(bold=True, color="FFFFFF", size=10)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border()


def thin_border() -> Border:
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def data_cell(cell, value=None, alt: bool = False, url: bool = False) -> None:
    if value is not None:
        cell.value = value
    cell.border = thin_border()
    cell.alignment = Alignment(vertical="center")
    if alt:
        cell.fill = PatternFill("solid", fgColor="EBF3FB")
    if url and cell.value:
        cell.hyperlink = cell.value
        cell.font = Font(color="0563C1", underline="single", size=9)
    else:
        cell.font = Font(size=9)


# ── Sheet 1: TM Links ─────────────────────────────────────────────────────────
def build_links_sheet(ws, players: list[dict]) -> None:
    ws.title = "TM Links"

    headers = ["No.", "Player Name", "Club", "TM ID", "Stats URL (click to open)"]
    col_widths = [6, 28, 22, 12, 70]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        header_style(cell)
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    for row_idx, p in enumerate(players, 2):
        alt = row_idx % 2 == 0
        values = [
            p["serial_no"],
            p["player_name"],
            p["club_name"],
            p["player_tm_id"],
            p["stats_url"],
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            data_cell(cell, alt=alt, url=(col == 5))
        ws.row_dimensions[row_idx].height = 15


# ── Sheet 2: Paste Data ───────────────────────────────────────────────────────
def build_paste_sheet(ws, players: list[dict]) -> None:
    ws.title = "Paste Data"

    headers = [
        "No.",
        "Player Name",
        "Club",
        "TM ID",
        "Season",  # e.g. 24/25
        "Competition",  # e.g. Premier League
        "Appearances",  # e.g. 38
        "Minutes Played",  # e.g. 3420
        # Extra cols in case you want to paste more
        "Col 2",
        "Col 3",
        "Yellow Cards",
        "2nd Yellow",
        "Red Cards",
        "Notes",
    ]
    col_widths = [6, 28, 22, 12, 8, 28, 13, 15, 10, 10, 10, 10, 10, 20]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        header_style(cell, bg="1A5276")
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "E2"  # freeze identity cols, scroll through season data

    # Pre-fill identity columns (No, Player Name, Club, TM ID) for all players
    # Leave season/competition/stats columns blank for you to paste into
    for row_idx, p in enumerate(players, 2):
        alt = row_idx % 2 == 0
        identity = [
            p["serial_no"],
            p["player_name"],
            p["club_name"],
            p["player_tm_id"],
        ]
        for col, val in enumerate(identity, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            data_cell(cell, alt=alt)

        # Leave cols 5-14 blank but styled
        for col in range(5, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col)
            data_cell(cell, alt=alt)

        ws.row_dimensions[row_idx].height = 15

    # Add a helper note at the top
    note_cell = ws.cell(
        row=1,
        column=6,
        value="← Paste season data here. One row per player per season per competition.",
    )
    note_cell.font = Font(italic=True, color="FFFFFF", size=9)


# ── Main ──────────────────────────────────────────────────────────────────────
def run(input_path: str, output_path: str) -> None:
    p = Path(input_path)
    df = (
        pd.read_excel(p, dtype=str)
        if p.suffix in (".xlsx", ".xls")
        else pd.read_csv(p, dtype=str)
    )
    df = df.fillna("")
    df.columns = [c.strip().lower().replace(" ", "_") for c in df.columns]
    df = df.drop_duplicates(subset=["player_tm_id"])
    print(f"Loaded {len(df)} unique players.")

    players = []
    for i, (_, row) in enumerate(df.iterrows(), 1):
        tm_id = str(row.get("player_tm_id", "")).strip()
        player_name = str(row.get("player_name", "")).strip()
        club_name = str(row.get("club_name", "")).strip()
        profile_url = str(row.get("profile_url", "")).strip()
        slug = extract_slug(profile_url)
        stats_url = build_stats_url(slug, tm_id)
        players.append(
            {
                "serial_no": i,
                "player_name": player_name,
                "club_name": club_name,
                "player_tm_id": tm_id,
                "stats_url": stats_url,
            }
        )

    wb = openpyxl.Workbook()

    # Sheet 1 — links (uses default sheet)
    ws1 = wb.active
    build_links_sheet(ws1, players)

    # Sheet 2 — paste area
    ws2 = wb.create_sheet()
    build_paste_sheet(ws2, players)

    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"✅ Saved to {out}")
    print(f"   Sheet 1 'TM Links'   — {len(players)} clickable links")
    print("   Sheet 2 'Paste Data' — ready for your manual data entry")
    print()
    print("Workflow:")
    print("  1. Open Sheet 1, click a player's Stats URL")
    print("  2. On TM, select all season rows (19/20 to 25/26)")
    print("  3. Go to Sheet 2, find that player's row, paste the data")
    print("  4. Repeat — one player at a time")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", "-i", default="data/raw/player_bios.csv")
    parser.add_argument("--output", "-o", default="data/raw/player_stats_links.xlsx")
    args = parser.parse_args()
    run(args.input, args.output)
