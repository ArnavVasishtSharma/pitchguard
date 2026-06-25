from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pathlib import Path

OUTPUT_FILE = r"C:\Users\Aarya\Downloads\Pitchguard\pitchguard\src\scrapers\data\raw\player_stats_links1.xlsx"

path = Path(OUTPUT_FILE)
print(f"File exists: {path.exists()}")

try:
    print("Opening file...")
    wb = load_workbook(path)
    ws = wb.active
    print(f"Max column currently: {ws.max_column}")
    print(f"Max row currently: {ws.max_row}")

    # Try writing one test cell
    next_col = (ws.max_column or 0) + 1
    col_letter = get_column_letter(next_col)
    print(f"Will write to column {next_col} ({col_letter})")
    ws[f"{col_letter}1"] = "TEST_VALUE"
    wb.save(OUTPUT_FILE)
    print("SUCCESS — write worked!")

    # Clean it up
    wb2 = load_workbook(path)
    ws2 = wb2.active
    ws2[f"{col_letter}1"] = None
    wb2.save(OUTPUT_FILE)
    print("Cleaned up test value.")

except PermissionError as e:
    print(f"PERMISSION ERROR: {e}")
    print("The file is open in Excel — close it first!")
except Exception as e:
    print(f"ERROR: {type(e).__name__}: {e}")
